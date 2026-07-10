import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


DARK_FILL = PatternFill("solid", fgColor="0D1117")
HEADER_FILL = PatternFill("solid", fgColor="00D4FF")
GREEN_FILL = PatternFill("solid", fgColor="00FF88")
RED_FILL = PatternFill("solid", fgColor="FF4444")
CARD_FILL = PatternFill("solid", fgColor="161B22")

WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
DARK_FONT = Font(color="0D1117", bold=True, name="Calibri", size=11)
BODY_FONT = Font(color="C9D1D9", name="Calibri", size=10)
TITLE_FONT = Font(color="00D4FF", bold=True, name="Calibri", size=16)

THIN = Side(style="thin", color="30363D")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_cell(cell, text):
    cell.value = text
    cell.font = DARK_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _style_body_cell(cell, value, number_format=None, fill=None):
    cell.value = value
    cell.font = BODY_FONT
    cell.fill = fill or CARD_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    if number_format:
        cell.number_format = number_format


def generate_excel(data: dict, path: str):
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00D4FF"

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 22
    for row in range(1, 60):
        ws.row_dimensions[row].height = 22

    ws["B2"].value = data["strategy_name"].upper()
    ws["B2"].font = TITLE_FONT
    ws["B2"].fill = DARK_FILL
    ws.merge_cells("B2:G2")

    ws["B3"].value = data["description"]
    ws["B3"].font = BODY_FONT
    ws["B3"].fill = DARK_FILL
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:G3")
    ws.row_dimensions[3].height = 50

    # Rules
    ws["B5"].value = "STRATEGY RULES"
    ws["B5"].font = WHITE_FONT
    ws["B5"].fill = PatternFill("solid", fgColor="161B22")
    ws.merge_cells("B5:G5")

    for i, rule in enumerate(data.get("rules", []), start=6):
        ws.cell(row=i, column=2).value = f"  • {rule}"
        ws.cell(row=i, column=2).font = BODY_FONT
        ws.cell(row=i, column=2).fill = DARK_FILL
        ws.merge_cells(f"B{i}:G{i}")

    row_offset = 6 + len(data.get("rules", []))

    # Metrics table
    m = data["metrics"]
    metrics_display = [
        ("Total Return", f"{m['total_return_pct']:+.2f}%"),
        ("CAGR", f"{m['cagr_pct']:.2f}%"),
        ("Sharpe Ratio", f"{m['sharpe_ratio']:.2f}"),
        ("Max Drawdown", f"{m['max_drawdown_pct']:.2f}%"),
        ("Win Rate", f"{m['win_rate_pct']:.1f}%"),
        ("Total Trades", str(m["total_trades"])),
        ("Profitable Trades", str(m["profitable_trades"])),
        ("Avg Trade Return", f"{m['avg_trade_return_pct']:+.2f}%"),
        ("Best Trade", f"{m['best_trade_pct']:+.2f}%"),
        ("Worst Trade", f"{m['worst_trade_pct']:.2f}%"),
        ("Starting Capital", f"₹{m['starting_capital']:,.0f}"),
        ("Ending Value", f"₹{m['ending_value']:,.0f}"),
    ]

    r = row_offset + 1
    _style_header_cell(ws.cell(row=r, column=2), "METRIC")
    _style_header_cell(ws.cell(row=r, column=4), "VALUE")
    ws.merge_cells(f"B{r}:C{r}")
    ws.merge_cells(f"D{r}:E{r}")

    for metric, value in metrics_display:
        r += 1
        cell_m = ws.cell(row=r, column=2)
        cell_v = ws.cell(row=r, column=4)
        _style_body_cell(cell_m, metric, fill=CARD_FILL)
        ws.merge_cells(f"B{r}:C{r}")
        color = GREEN_FILL if ("+" in value or (value.replace("₹","").replace(",","").replace(".","").replace("%","").lstrip("-").isdigit() and float(value.replace("₹","").replace(",","").replace("%","").replace("+","")) > 0)) else CARD_FILL
        _style_body_cell(cell_v, value, fill=color)
        ws.merge_cells(f"D{r}:E{r}")

    # ── Sheet 2: Trade Log ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Trade Log")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "00FF88"

    headers = ["Date", "Action", "Price (₹)", "Units", "P&L (₹)", "Cumulative P&L (₹)"]
    col_widths = [15, 10, 15, 10, 18, 22]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=2):
        ws2.column_dimensions[get_column_letter(i)].width = w
        _style_header_cell(ws2.cell(row=2, column=i), h)

    for r_idx, trade in enumerate(data.get("trades", []), start=3):
        pnl = trade.get("pnl", 0)
        cum_pnl = trade.get("cumulative_pnl", 0)
        row_fill = GREEN_FILL if pnl >= 0 else RED_FILL
        values = [
            trade["date"],
            trade["action"],
            trade["price"],
            trade.get("units", 0),
            pnl,
            cum_pnl,
        ]
        for c_idx, val in enumerate(values, start=2):
            cell = ws2.cell(row=r_idx, column=c_idx)
            fmt = None
            if c_idx in (3, 5, 6):
                fmt = '#,##0.00'
            _style_body_cell(cell, val, number_format=fmt,
                             fill=row_fill if c_idx in (5, 6) else CARD_FILL)

    # ── Sheet 3: Equity Curve ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Equity Curve")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "FF8800"

    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 60

    _style_header_cell(ws3.cell(row=2, column=2), "Date")
    _style_header_cell(ws3.cell(row=2, column=3), "Portfolio Value (₹)")
    _style_header_cell(ws3.cell(row=2, column=4), "Month")
    _style_header_cell(ws3.cell(row=2, column=5), "Monthly Return (%)")

    equity = data.get("equity_curve", [])
    monthly = data.get("monthly_returns", [])

    for i, eq in enumerate(equity, start=3):
        ws3.cell(row=i, column=2).value = eq["date"]
        ws3.cell(row=i, column=2).font = BODY_FONT
        ws3.cell(row=i, column=2).fill = CARD_FILL
        _style_body_cell(ws3.cell(row=i, column=3), eq["portfolio_value"], '#,##0.00', CARD_FILL)

    for i, mr in enumerate(monthly, start=3):
        ws3.cell(row=i, column=4).value = mr["month"]
        ws3.cell(row=i, column=4).font = BODY_FONT
        ws3.cell(row=i, column=4).fill = CARD_FILL
        ret = mr["return_pct"]
        fill = GREEN_FILL if ret >= 0 else RED_FILL
        _style_body_cell(ws3.cell(row=i, column=5), ret, '0.00"%"', fill)

    # Add line chart
    if equity:
        chart = LineChart()
        chart.title = "Portfolio Equity Curve"
        chart.style = 10
        chart.y_axis.title = "Portfolio Value (₹)"
        chart.x_axis.title = "Date"
        chart.height = 15
        chart.width = 30

        end_row = 2 + len(equity)
        data_ref = Reference(ws3, min_col=3, min_row=2, max_row=end_row)
        chart.add_data(data_ref, titles_from_data=True)

        ws3.add_chart(chart, "G2")

    wb.save(path)


def generate_excel_to_path(data: dict, download_dir: str) -> str:
    os.makedirs(download_dir, exist_ok=True)
    path = os.path.join(download_dir, "backtest_report.xlsx")
    generate_excel(data, path)
    return path
