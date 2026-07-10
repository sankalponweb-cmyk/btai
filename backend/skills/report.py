"""
report.py
=========
Build the polished, multi-sheet Excel report from backtest results.

Sheets produced (in this order):
1. Strategy Performance  - headline metrics, MOSTLY FORMULAS referencing
                            the 'Equity & Drawdown' sheet, so the report
                            stays correct if that data is edited/extended.
2. Trade Log              - one row per (partial) trade, raw simulation
                            output, plus a few helper columns (win flag,
                            win/loss streaks) used by Trade Statistics.
3. Historical Data        - OHLCV data for every ticker used, stacked with
                            a Ticker column.
4. Equity & Drawdown      - daily Date/Equity/Benchmark/Exposure (raw
                            outputs) plus Daily Return/Drawdown/Underwater
                            Streak/Downside Return (formulas), with a combo
                            line chart (equity + benchmark on the primary
                            axis, drawdown % on the secondary axis).
5. Trade Statistics       - trade-level stats, MOSTLY FORMULAS referencing
                            the Trade Log sheet (COUNTIF/AVERAGEIF/SUMIF),
                            plus a per-instrument breakdown and an annual
                            returns table.

Only `build_report(...)` needs to be called; the helper functions are kept
top-level for readability/testability.

IMPORTANT: after writing the .xlsx, run scripts/recalc.py (from the xlsx
skill) on the output file so the formula results are populated and any
formula errors are caught before handing the file to the user.
"""

import io
import zipfile

import numpy as np
import pandas as pd
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, AreaChart, Reference
from openpyxl.chart.legend import Legend

# DrawingML namespace used inside chart spPr elements
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"

# --------------------------------------------------------------------------- #
# Styling constants
# --------------------------------------------------------------------------- #
FONT_NAME = "Arial"

TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
LABEL_FONT = Font(name=FONT_NAME, bold=True)
NORMAL_FONT = Font(name=FONT_NAME)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF")  # blue = editable input

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HELPER_FILL = PatternFill("solid", start_color="F2F2F2")
ALT_FILL = PatternFill("solid", start_color="F8FAFC")

THIN = Side(border_style="thin", color="D9D9D9")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT = "0.00%"
NUM_FMT = "#,##0.00"
INT_FMT = "#,##0"
RATIO_FMT = "0.00"
DATE_FMT = "yyyy-mm-dd"

# Fixed cell address (on the Strategy Performance sheet) for the risk-free
# rate input. Other sheets reference this directly, so the Strategy
# Performance sheet layout (build_strategy_performance_sheet) uses a FIXED
# number of header rows -- regardless of whether optional fields like
# description/universe are provided -- so this address never drifts.
RF_CELL = "'Strategy Performance'!$C$16"


# --------------------------------------------------------------------------- #
# Generic helpers
# --------------------------------------------------------------------------- #
def _write_header_row(ws, row, headers, start_col=1, fill=HEADER_FILL, font=HEADER_FONT):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, text, row=1, span=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _kv_row(ws, row, label, value, value_fmt=None, is_input=False, note=None):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT
    vc = ws.cell(row=row, column=3, value=value)
    vc.font = INPUT_FONT if is_input else NORMAL_FONT
    if value_fmt:
        vc.number_format = value_fmt
    if note:
        nc = ws.cell(row=row, column=5, value=note)
        nc.font = Font(name=FONT_NAME, italic=True, size=9, color="808080")
    return vc


# --------------------------------------------------------------------------- #
# Sheet 4: Equity & Drawdown  (built first so other sheets can reference it)
# --------------------------------------------------------------------------- #
def build_equity_drawdown_sheet(wb, result, periods_per_year=252):
    ws = wb.create_sheet("Equity & Drawdown")
    eq = result.equity_curve.dropna()
    exposure = result.exposure.reindex(eq.index) if result.exposure is not None else pd.Series(index=eq.index, dtype=float)
    benchmark = None
    if result.benchmark_curve is not None:
        benchmark = result.benchmark_curve.reindex(eq.index)

    has_benchmark = benchmark is not None and benchmark.notna().any()

    headers = ["Date", "Equity", "Benchmark (Buy & Hold)", "Exposure (%)",
               "Daily Return", "Drawdown (%)", "Underwater Streak (bars)",
               "Downside Excess Return"]
    _write_header_row(ws, 1, headers)

    n = len(eq)
    first_data_row = 2
    last_row = first_data_row + n - 1

    for i, (date, equity_val) in enumerate(eq.items()):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=date.to_pydatetime()).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=float(equity_val)).number_format = NUM_FMT

        bench_val = None
        if has_benchmark:
            bv = benchmark.iloc[i]
            if pd.notna(bv):
                bench_val = float(bv)
        c3 = ws.cell(row=r, column=3, value=bench_val)
        c3.number_format = NUM_FMT

        exp_val = exposure.iloc[i] if i < len(exposure) and pd.notna(exposure.iloc[i]) else 0.0
        ws.cell(row=r, column=4, value=float(exp_val)).number_format = PCT_FMT

        # Daily Return (formula): blank/0 for the first row
        if i == 0:
            ws.cell(row=r, column=5, value=0).number_format = PCT_FMT
        else:
            ws.cell(row=r, column=5, value=f"=(B{r}-B{r-1})/B{r-1}").number_format = PCT_FMT

        # Drawdown % = equity / running max equity - 1
        ws.cell(row=r, column=6, value=f"=B{r}/MAX($B${first_data_row}:B{r})-1").number_format = PCT_FMT

        # Underwater streak (consecutive bars with drawdown < 0)
        if i == 0:
            ws.cell(row=r, column=7, value=f"=IF(F{r}<0,1,0)")
        else:
            ws.cell(row=r, column=7, value=f"=IF(F{r}<0,G{r-1}+1,0)")
        ws.cell(row=r, column=7).number_format = INT_FMT

        # Downside excess return (helper for Sortino): negative (return - rf_per_period) only
        ws.cell(
            row=r, column=8,
            value=f'=IF((E{r}-{RF_CELL}/{periods_per_year})<0,(E{r}-{RF_CELL}/{periods_per_year}),"")'
        ).number_format = PCT_FMT
        ws.cell(row=r, column=8).fill = HELPER_FILL

    ws.cell(row=1, column=8).fill = HEADER_FILL  # keep header style on helper col

    _set_col_widths(ws, [12, 16, 20, 14, 14, 14, 18, 18])
    ws.freeze_panes = "A2"

    # ── Combo chart ─────────────────────────────────────────────────────────
    # Equity + Benchmark: line series on the primary (left) Y-axis
    # Drawdown:           filled area on the secondary (right) Y-axis
    # ─────────────────────────────────────────────────────────────────────────

    # Primary line chart — equity (blue) and benchmark (dark red)
    chart = LineChart()
    chart.title = "Equity Curve & Drawdown"
    chart.style = 2
    chart.height = 15   # cm
    chart.width  = 30   # cm

    # ── Primary Y-axis (left): portfolio equity ──────────────────────────────
    chart.y_axis.axId    = 100
    chart.y_axis.title   = "Equity (₹)"
    chart.y_axis.numFmt  = "#,##0"
    chart.y_axis.crossAx = 10     # paired with the X-axis
    chart.y_axis.delete  = False

    # ── X-axis (dates): category axis using the Date column ──────────────────
    # Note: majorTimeUnit is a DateAx-only property and must NOT be set on a
    # CatAx (which LineChart uses); it would be silently ignored at best or
    # produce blank tick labels at worst.  The cell format (yyyy-mm-dd) drives
    # how the category labels are rendered.
    chart.x_axis.axId    = 10
    chart.x_axis.title   = "Date"
    chart.x_axis.numFmt  = "yyyy-mm-dd"   # must use numFmt, not number_format
    chart.x_axis.crossAx = 100    # paired with the primary Y-axis
    chart.x_axis.delete  = False

    # Equity series
    eq_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=last_row)
    chart.add_data(eq_ref, titles_from_data=True)
    s_eq = chart.series[0]
    s_eq.graphicalProperties.line.solidFill = "2E75B6"   # blue
    s_eq.graphicalProperties.line.width = 19050           # 1.5 pt
    s_eq.smooth = False
    s_eq.marker.symbol = "none"

    # Benchmark series (only if benchmark data is present)
    if has_benchmark:
        bench_ref = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=last_row)
        chart.add_data(bench_ref, titles_from_data=True)
        s_bm = chart.series[1]
        s_bm.graphicalProperties.line.solidFill = "C00000"  # dark red
        s_bm.graphicalProperties.line.width = 12700          # 1 pt
        s_bm.smooth = False
        s_bm.marker.symbol = "none"

    # X-axis categories (shared with drawdown area)
    dates_ref = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.set_categories(dates_ref)

    # Legend at bottom, matching the screenshot layout
    lgd = Legend()
    lgd.position = "b"
    chart.legend = lgd

    # Drawdown area chart — secondary (right) Y-axis
    # crosses="max" puts 0 % at the top so the area fills downward
    dd_chart = AreaChart()
    dd_chart.grouping = "standard"
    dd_ref = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=last_row)
    dd_chart.add_data(dd_ref, titles_from_data=True)

    s_dd = dd_chart.series[0]
    s_dd.graphicalProperties.solidFill = "A9D18E"          # light sage green fill
    s_dd.graphicalProperties.line.solidFill = "70AD47"     # slightly darker green border
    s_dd.graphicalProperties.line.width = 6350              # 0.5 pt border line

    # ── Secondary Y-axis (right): drawdown % ─────────────────────────────────
    dd_chart.y_axis.axId    = 200
    dd_chart.y_axis.title   = "Drawdown (%)"
    dd_chart.y_axis.numFmt  = "0%"
    dd_chart.y_axis.crossAx = 10     # paired with the same X-axis (axId 10)
    dd_chart.y_axis.crosses = "max"  # 0 % at top; area fills downward
    dd_chart.y_axis.delete  = False

    # AreaChart gets its own x-axis object; mark it deleted so only one set of
    # date labels appears at the bottom (the primary LineChart's x-axis).
    dd_chart.x_axis.axId    = 10
    dd_chart.x_axis.delete  = True

    # Merge into one chart object; area chart draws behind the line series
    chart += dd_chart

    ws.add_chart(chart, "J2")

    return {
        "sheet": "Equity & Drawdown",
        "first_row": first_data_row,
        "last_row": last_row,
        "has_benchmark": has_benchmark,
    }


# --------------------------------------------------------------------------- #
# Sheet 2: Trade Log
# --------------------------------------------------------------------------- #
def build_trade_log_sheet(wb, trade_log_df):
    ws = wb.create_sheet("Trade Log")

    headers = ["Ticker", "Side", "Entry Date", "Entry Price", "Exit Date",
               "Exit Price", "Shares", "Holding Days", "P&L", "Return (%)",
               "Exit Reason", "Win (1/0)", "Win Streak", "Loss Streak"]
    _write_header_row(ws, 1, headers)

    if trade_log_df is None or trade_log_df.empty:
        ws.cell(row=2, column=1, value="No trades were executed during the backtest period.").font = NORMAL_FONT
        _set_col_widths(ws, [12, 8, 12, 12, 12, 12, 10, 12, 14, 12, 16, 10, 10, 10])
        return {"sheet": "Trade Log", "first_row": 2, "last_row": 1, "has_trades": False}

    df = trade_log_df.reset_index(drop=True)
    first_data_row = 2
    last_row = first_data_row + len(df) - 1

    for i, row in df.iterrows():
        r = first_data_row + i
        ws.cell(row=r, column=1, value=str(row["ticker"]))
        ws.cell(row=r, column=2, value=str(row["side"]))
        ws.cell(row=r, column=3, value=pd.Timestamp(row["entry_date"]).to_pydatetime()).number_format = DATE_FMT
        ws.cell(row=r, column=4, value=float(row["entry_price"])).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=pd.Timestamp(row["exit_date"]).to_pydatetime()).number_format = DATE_FMT
        ws.cell(row=r, column=6, value=float(row["exit_price"])).number_format = NUM_FMT
        ws.cell(row=r, column=7, value=float(row["shares"])).number_format = NUM_FMT
        ws.cell(row=r, column=8, value=float(row["holding_days"])).number_format = INT_FMT
        pnl_cell = ws.cell(row=r, column=9, value=float(row["pnl"]))
        pnl_cell.number_format = NUM_FMT
        pnl_cell.font = Font(name=FONT_NAME, color="008000" if row["pnl"] > 0 else "FF0000")
        ws.cell(row=r, column=10, value=float(row["return_pct"])).number_format = PCT_FMT
        ws.cell(row=r, column=11, value=str(row.get("exit_reason", "")))

        # Helper columns (formulas)
        win_cell = ws.cell(row=r, column=12, value=f"=IF(I{r}>0,1,0)")
        win_cell.fill = HELPER_FILL
        win_cell.number_format = INT_FMT

        if i == 0:
            ws.cell(row=r, column=13, value=f"=IF(L{r}=1,1,0)")
            ws.cell(row=r, column=14, value=f"=IF(L{r}=0,1,0)")
        else:
            ws.cell(row=r, column=13, value=f"=IF(L{r}=1,M{r-1}+1,0)")
            ws.cell(row=r, column=14, value=f"=IF(L{r}=0,N{r-1}+1,0)")
        for col in (13, 14):
            cell = ws.cell(row=r, column=col)
            cell.fill = HELPER_FILL
            cell.number_format = INT_FMT

    ws.cell(row=1, column=12).fill = HEADER_FILL
    ws.cell(row=1, column=13).fill = HEADER_FILL
    ws.cell(row=1, column=14).fill = HEADER_FILL

    _set_col_widths(ws, [12, 8, 12, 12, 12, 12, 10, 12, 14, 12, 16, 10, 10, 10])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{last_row}"

    return {"sheet": "Trade Log", "first_row": first_data_row, "last_row": last_row, "has_trades": True}


# --------------------------------------------------------------------------- #
# Sheet 3: Historical Data
# --------------------------------------------------------------------------- #
def build_historical_data_sheet(wb, price_data):
    ws = wb.create_sheet("Historical Data")
    headers = ["Ticker", "Date", "Open", "High", "Low", "Close", "Volume"]
    _write_header_row(ws, 1, headers)

    r = 2
    for ticker, df in price_data.items():
        d = df.copy()
        d.columns = [str(c).strip().lower() for c in d.columns]
        d.index = pd.to_datetime(d.index)
        d = d.sort_index()
        for date, row in d.iterrows():
            ws.cell(row=r, column=1, value=str(ticker))
            ws.cell(row=r, column=2, value=date.to_pydatetime()).number_format = DATE_FMT
            for j, col in enumerate(["open", "high", "low", "close"], start=3):
                val = row.get(col)
                c = ws.cell(row=r, column=j, value=float(val) if pd.notna(val) else None)
                c.number_format = NUM_FMT
            vol = row.get("volume")
            ws.cell(row=r, column=7, value=int(vol) if pd.notna(vol) else None).number_format = INT_FMT
            r += 1

    last_row = r - 1
    _set_col_widths(ws, [14, 12, 12, 12, 12, 12, 14])
    ws.freeze_panes = "A2"
    if last_row >= 1:
        ws.auto_filter.ref = f"A1:G{max(last_row, 1)}"

    return {"sheet": "Historical Data", "first_row": 2, "last_row": last_row}


# --------------------------------------------------------------------------- #
# Sheet 5: Trade Statistics
# --------------------------------------------------------------------------- #
def build_trade_statistics_sheet(wb, trade_log_meta, tickers, annual_returns_df=None):
    ws = wb.create_sheet("Trade Statistics")
    _title(ws, "Trade Statistics", row=1, span=4)

    has_trades = trade_log_meta.get("has_trades", False)
    fr, lr = trade_log_meta["first_row"], trade_log_meta["last_row"]
    TL = "'Trade Log'"
    pnl = f"{TL}!I{fr}:I{lr}"
    ret = f"{TL}!J{fr}:J{lr}"
    hold = f"{TL}!H{fr}:H{lr}"
    win_streak = f"{TL}!M{fr}:M{lr}"
    loss_streak = f"{TL}!N{fr}:N{lr}"
    tick_col = f"{TL}!A{fr}:A{lr}"

    row = 3
    _write_header_row(ws, row, ["Metric", "", "Value"], start_col=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    if not has_trades:
        ws.cell(row=row, column=1, value="No trades were executed during the backtest period.")
        _set_col_widths(ws, [30, 4, 16, 4])
        return

    def kv(label, formula, fmt=NUM_FMT):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=3, value=formula)
        c.number_format = fmt
        row += 1

    kv("Total Trades", f"=COUNTA({tick_col})", INT_FMT)
    kv("Winning Trades", f'=COUNTIF({pnl},">0")', INT_FMT)
    kv("Losing Trades", f'=COUNTIF({pnl},"<=0")', INT_FMT)
    kv("Win Rate", f'=IFERROR(COUNTIF({pnl},">0")/COUNTA({tick_col}),0)', PCT_FMT)
    kv("Gross Profit", f'=SUMIF({pnl},">0")', NUM_FMT)
    kv("Gross Loss", f'=SUMIF({pnl},"<=0")', NUM_FMT)
    kv("Profit Factor", f'=IF(SUMIF({pnl},"<=0")=0,"n/a (no losses)",'
                         f'SUMIF({pnl},">0")/ABS(SUMIF({pnl},"<=0")))', RATIO_FMT)
    kv("Avg Win", f'=IFERROR(AVERAGEIF({pnl},">0"),0)', NUM_FMT)
    kv("Avg Loss", f'=IFERROR(AVERAGEIF({pnl},"<=0"),0)', NUM_FMT)
    kv("Payoff Ratio", f'=IF(IFERROR(AVERAGEIF({pnl},"<=0"),0)=0,"n/a (no losses)",'
                        f'IFERROR(AVERAGEIF({pnl},">0"),0)/ABS(AVERAGEIF({pnl},"<=0")))', RATIO_FMT)
    kv("Expectancy (per trade)", f"=AVERAGE({pnl})", NUM_FMT)
    kv("Avg Trade Return", f"=AVERAGE({ret})", PCT_FMT)
    kv("Avg Win Return", f'=IFERROR(AVERAGEIF({pnl},">0",{ret}),0)', PCT_FMT)
    kv("Avg Loss Return", f'=IFERROR(AVERAGEIF({pnl},"<=0",{ret}),0)', PCT_FMT)
    kv("Largest Win", f"=MAX({pnl})", NUM_FMT)
    kv("Largest Loss", f"=MIN({pnl})", NUM_FMT)
    kv("Avg Holding Period (days)", f"=AVERAGE({hold})", RATIO_FMT)
    kv("Avg Holding Period - Wins (days)", f'=IFERROR(AVERAGEIF({pnl},">0",{hold}),0)', RATIO_FMT)
    kv("Avg Holding Period - Losses (days)", f'=IFERROR(AVERAGEIF({pnl},"<=0",{hold}),0)', RATIO_FMT)
    kv("Max Consecutive Wins", f"=MAX({win_streak})", INT_FMT)
    kv("Max Consecutive Losses", f"=MAX({loss_streak})", INT_FMT)

    row += 1

    # -- Breakdown by instrument (only meaningful for multi-ticker portfolios)
    if tickers and len(tickers) > 1:
        row += 1
        ws.cell(row=row, column=1, value="Breakdown by Instrument").font = SUBTITLE_FONT
        row += 1
        _write_header_row(ws, row, ["Ticker", "Trades", "Win Rate", "Total P&L", "Avg Return"], start_col=1)
        row += 1
        for t in tickers:
            ws.cell(row=row, column=1, value=t)
            ws.cell(row=row, column=2,
                    value=f'=COUNTIF({tick_col},$A{row})').number_format = INT_FMT
            ws.cell(row=row, column=3,
                    value=f'=IFERROR(COUNTIFS({tick_col},$A{row},{pnl},">0")/COUNTIF({tick_col},$A{row}),0)'
                    ).number_format = PCT_FMT
            ws.cell(row=row, column=4,
                    value=f'=SUMIF({tick_col},$A{row},{pnl})').number_format = NUM_FMT
            ws.cell(row=row, column=5,
                    value=f'=IFERROR(AVERAGEIF({tick_col},$A{row},{ret}),0)').number_format = PCT_FMT
            row += 1

    # -- Annual returns table (computed in Python from the equity curve)
    if annual_returns_df is not None and not annual_returns_df.empty:
        row += 1
        ws.cell(row=row, column=1, value="Annual Returns").font = SUBTITLE_FONT
        row += 1
        _write_header_row(ws, row, ["Year", "End Equity", "Return (%)"], start_col=1)
        row += 1
        for _, r in annual_returns_df.iterrows():
            ws.cell(row=row, column=1, value=int(r["Year"]))
            ws.cell(row=row, column=2, value=float(r["End Equity"])).number_format = NUM_FMT
            ws.cell(row=row, column=3, value=float(r["Return (%)"]) / 100).number_format = PCT_FMT
            row += 1

    _set_col_widths(ws, [32, 16, 14, 16, 14])


# --------------------------------------------------------------------------- #
# Sheet 1: Strategy Performance
# --------------------------------------------------------------------------- #
def build_strategy_performance_sheet(wb, eq_meta, strategy_info, risk_free_rate, periods_per_year=252):
    ws = wb.create_sheet("Strategy Performance", 0)  # make it the first sheet
    _title(ws, strategy_info.get("name", "Strategy Performance"), row=1, span=6)

    # Fixed-position header rows (rows 2-5) regardless of which optional
    # fields are supplied -- this keeps the row numbers below, and the
    # RF_CELL constant referenced from other sheets, deterministic.
    desc = strategy_info.get("description", "")
    c = ws.cell(row=2, column=1, value=desc)
    c.font = Font(name=FONT_NAME, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 32

    universe = strategy_info.get("universe", [])
    ws.cell(row=3, column=1, value="Universe").font = LABEL_FONT
    ws.cell(row=3, column=3, value=", ".join(universe) if universe else "").font = NORMAL_FONT
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=6)

    ws.cell(row=4, column=1, value="Commission (bps, per side)").font = LABEL_FONT
    ws.cell(row=4, column=3, value=strategy_info.get("commission_bps", 0)).number_format = RATIO_FMT

    ws.cell(row=5, column=1, value="Slippage (bps, per side)").font = LABEL_FONT
    ws.cell(row=5, column=3, value=strategy_info.get("slippage_bps", 0)).number_format = RATIO_FMT

    row = 7
    _write_header_row(ws, row, ["Metric", "", "Value"], start_col=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    EQ = "'Equity & Drawdown'"
    fr, lr = eq_meta["first_row"], eq_meta["last_row"]
    ret_range = f"{EQ}!E{fr}:E{lr}"
    dd_range = f"{EQ}!F{fr}:F{lr}"
    streak_range = f"{EQ}!G{fr}:G{lr}"
    downside_range = f"{EQ}!H{fr}:H{lr}"
    date_first = f"{EQ}!A{fr}"
    date_last = f"{EQ}!A{lr}"

    def kv(label, formula_or_value, fmt=NUM_FMT, is_input=False, note=None):
        nonlocal row
        _kv_row(ws, row, label, formula_or_value, value_fmt=fmt, is_input=is_input, note=note)
        row += 1

    kv("Start Date", f"={date_first}", DATE_FMT)
    start_date_row = row - 1
    kv("End Date", f"={date_last}", DATE_FMT)
    end_date_row = row - 1

    kv("Duration (Years)", f"=(C{end_date_row}-C{start_date_row})/365.25", RATIO_FMT)
    duration_row = row - 1

    kv("Initial Capital", f"={EQ}!B{fr}", NUM_FMT)
    initial_capital_row = row - 1
    kv("Final Equity", f"={EQ}!B{lr}", NUM_FMT)
    final_equity_row = row - 1

    kv("Total Return", f"=C{final_equity_row}/C{initial_capital_row}-1", PCT_FMT)

    kv("CAGR", f"=(C{final_equity_row}/C{initial_capital_row})^(1/C{duration_row})-1", PCT_FMT)
    cagr_row = row - 1

    kv("Annualized Volatility", f"=STDEV({ret_range})*SQRT({periods_per_year})", PCT_FMT)

    rf_row = row
    kv("Risk-Free Rate (annual, editable)", risk_free_rate, PCT_FMT, is_input=True,
       note="Used for Sharpe / Sortino excess return")
    assert f"$C${rf_row}" in RF_CELL, "RF_CELL constant must match this row"

    kv("Sharpe Ratio",
       f"=IFERROR((AVERAGE({ret_range})-C{rf_row}/{periods_per_year})"
       f"/STDEV({ret_range})*SQRT({periods_per_year}),0)", RATIO_FMT)

    kv("Sortino Ratio",
       f"=IFERROR((AVERAGE({ret_range})-C{rf_row}/{periods_per_year})"
       f"/STDEV({downside_range})*SQRT({periods_per_year}),0)", RATIO_FMT)

    kv("Max Drawdown", f"=MIN({dd_range})", PCT_FMT)
    max_dd_row = row - 1

    kv("Max Drawdown Duration (bars)", f"=MAX({streak_range})", INT_FMT)

    kv("Calmar Ratio", f"=IFERROR(C{cagr_row}/ABS(C{max_dd_row}),0)", RATIO_FMT)

    kv("Average Exposure", f"=AVERAGE({EQ}!D{fr}:D{lr})", PCT_FMT)
    kv("Best Day", f"=MAX({ret_range})", PCT_FMT)
    kv("Worst Day", f"=MIN({ret_range})", PCT_FMT)
    kv("Positive Days", f'=COUNTIF({ret_range},">0")/COUNT({ret_range})', PCT_FMT)

    if eq_meta.get("has_benchmark"):
        bench_first = f"{EQ}!C{fr}"
        bench_last = f"{EQ}!C{lr}"
        kv("Benchmark Total Return", f"={bench_last}/{bench_first}-1", PCT_FMT)
        kv("Benchmark CAGR", f"=(({bench_last}/{bench_first}))^(1/C{duration_row})-1", PCT_FMT)

    _set_col_widths(ws, [34, 2, 16, 4, 30, 4])


# --------------------------------------------------------------------------- #
# Post-save gradient patch
# --------------------------------------------------------------------------- #
def _patch_drawdown_gradient(xlsx_path: str,
                              hex_top: str = "70AD47",
                              hex_bottom: str = "E2EFDA") -> None:
    """
    openpyxl's chart-series API only exposes solidFill; there is no clean way
    to set a gradient fill on a chart series before serialisation.

    This function opens the saved xlsx (which is a ZIP of XML files), locates
    every <c:areaChart><c:ser><c:spPr> element, replaces the <a:solidFill>
    with a top-to-bottom linear <a:gradFill>, then writes the file back.

    hex_top    : colour at pos=0   (top of plot area = 0 % drawdown line)
    hex_bottom : colour at pos=100000 (bottom of plot area = max drawdown)
    """
    try:
        # Read the xlsx ZIP into memory
        with zipfile.ZipFile(xlsx_path, "r") as zin:
            all_names = zin.namelist()
            contents  = {n: zin.read(n) for n in all_names}

        chart_files = [n for n in all_names
                       if n.startswith("xl/charts/chart") and n.endswith(".xml")]

        modified = False
        for cf in chart_files:
            tree = etree.fromstring(contents[cf])

            # All area-chart series shape-property containers
            spPrs = tree.findall(
                f".//{{{_C_NS}}}areaChart/{{{_C_NS}}}ser/{{{_C_NS}}}spPr"
            )
            if not spPrs:
                continue

            for spPr in spPrs:
                # Remove any existing fill element
                for fill_tag in ("solidFill", "gradFill", "noFill", "pattFill"):
                    for el in spPr.findall(f"{{{_A_NS}}}{fill_tag}"):
                        spPr.remove(el)

                # Build the gradient fill element
                gf    = etree.Element(f"{{{_A_NS}}}gradFill")
                gsLst = etree.SubElement(gf, f"{{{_A_NS}}}gsLst")

                gs0 = etree.SubElement(gsLst, f"{{{_A_NS}}}gs", attrib={"pos": "0"})
                etree.SubElement(gs0, f"{{{_A_NS}}}srgbClr", attrib={"val": hex_top})

                gs1 = etree.SubElement(gsLst, f"{{{_A_NS}}}gs", attrib={"pos": "100000"})
                etree.SubElement(gs1, f"{{{_A_NS}}}srgbClr", attrib={"val": hex_bottom})

                # ang=5 400 000 = 90° in 60 000ths-of-a-degree → top-to-bottom
                etree.SubElement(gf, f"{{{_A_NS}}}lin",
                                 attrib={"ang": "5400000", "scaled": "0"})

                # Insert before <a:ln> so DrawingML element order is valid
                ln_el = spPr.find(f"{{{_A_NS}}}ln")
                if ln_el is not None:
                    children = list(spPr)
                    spPr.insert(children.index(ln_el), gf)
                else:
                    spPr.append(gf)

            contents[cf] = etree.tostring(
                tree, xml_declaration=True, encoding="UTF-8", standalone=True
            )
            modified = True

        if modified:
            with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name, data in contents.items():
                    zout.writestr(name, data)

    except Exception as exc:   # non-fatal: report is usable without gradient
        import sys
        print(f"[report] gradient patch skipped: {exc}", file=sys.stderr)


# --------------------------------------------------------------------------- #
# Orchestrator
# --------------------------------------------------------------------------- #
def build_report(output_path, result, price_data, strategy_info=None,
                  risk_free_rate=0.0, periods_per_year=252, annual_returns_df=None):
    """
    Parameters
    ----------
    output_path : str
        Path to write the .xlsx file to.
    result : engine.BacktestResult
    price_data : dict[str, pd.DataFrame]
        The raw historical OHLCV data downloaded for the backtest (used to
        populate the 'Historical Data' sheet).
    strategy_info : dict, optional
        {"name": str, "description": str, "universe": list[str],
         "commission_bps": float, "slippage_bps": float}
    risk_free_rate : float
        Annual risk-free rate as a decimal (e.g. 0.07 for 7%), written as an
        editable input cell that drives Sharpe/Sortino.
    periods_per_year : int
        252 for daily data, 52 for weekly, 12 for monthly.
    annual_returns_df : pd.DataFrame, optional
        Output of metrics.annual_returns(result.equity_curve).
    """
    strategy_info = strategy_info or {}
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    eq_meta = build_equity_drawdown_sheet(wb, result, periods_per_year=periods_per_year)
    trade_meta = build_trade_log_sheet(wb, result.trade_log)
    build_historical_data_sheet(wb, price_data)
    build_trade_statistics_sheet(wb, trade_meta, result.meta.get("tickers", []),
                                  annual_returns_df=annual_returns_df)
    build_strategy_performance_sheet(wb, eq_meta, strategy_info, risk_free_rate,
                                      periods_per_year=periods_per_year)

    # Reorder sheets into the desired final order
    order = ["Strategy Performance", "Trade Log", "Historical Data", "Equity & Drawdown", "Trade Statistics"]
    wb._sheets.sort(key=lambda s: order.index(s.title))

    wb.save(output_path)

    # openpyxl can't set gradient fills on chart series before serialisation,
    # so patch the saved file's chart XML directly.
    _patch_drawdown_gradient(output_path)

    return output_path
